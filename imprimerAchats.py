import os
import csv
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from tkinter import Tk, filedialog
from datetime import date
from reportlab.lib.units import inch 
from openpyxl import Workbook #POUR CREE UN CLASSEUR EXCEL


# Chemin du fichier CSV
chemin_fichier_csv = 'achats.csv'

# Chemin du fichier Excel de sortie
chemin_fichier_excel = 'donneesAchats.xlsx'

# Création d'un nouveau classeur Excel
classeur = Workbook()
feuille = classeur.active

# Lecture des données depuis le fichier CSV et écriture dans le classeur Excel
with open(chemin_fichier_csv, 'r') as fichier_csv:
    lecteur_csv = csv.reader(fichier_csv)
    for ligne in lecteur_csv:
        feuille.append(ligne)

# Enregistrement du classeur Excel
classeur.save(chemin_fichier_excel)


def choisir_emplacement_enregistrement():
    """Fonction pour choisir l'emplacement où enregistrer le fichier PDF."""
    root = Tk()
    root.attributes('-topmost', True)
    root.withdraw()
    chemin_dossier = filedialog.askdirectory()
    root.destroy()
    return chemin_dossier

def creer_facture(data, chemin_sortie):
    # Création du PDF
    nom_fichier_pdf = os.path.join(chemin_sortie, "factureAchats.pdf")
    c = canvas.Canvas(nom_fichier_pdf, pagesize=letter)
    
# Calcul des coordonnées de la page
    page_width, page_height = letter
    x1, y1 = 0, 0
    x2, y2 = page_width, page_height

# Dessiner un cadre autour de la page
    c.setStrokeColorRGB(0, 0, 0)  # Couleur du trait noir
    c.rect(x1+30, y1+30, x2-50, y2-50)

    # Définir les coordonnées du titre et les dimensions du rectangle
    titre_x = 50
    titre_y = 730
    titre_largeur = 500
    titre_hauteur = 20
    # Écriture des données dans le PDF
    c.setFont("Helvetica",16)
    c.setFillColorRGB(0.5, 0.1, 0.1)  # BORDEAU
# Dessiner un rectangle autour du titre
    c.rect(titre_x, titre_y, titre_largeur, titre_hauteur)
    # Écrire le titre à l'intérieur du rectangle encadré
    c.setFont("Helvetica", 12)
    c.drawString(titre_x + 5, titre_y + 5, "FACTURE")
    # Ajout de la date d'aujourd'hui dans le PDF

    c.setFillColorRGB(0, 0, 0)  # Noir

    c.setFont("Helvetica", 10)
    aujourdhui = date.today().strftime("%d/%m/%Y")
    c.drawString(460, 740, f"Date: {aujourdhui}")
    
    # Informations de l'entreprise
    c.setFont("Helvetica",10)
    nom_entreprise = "Nom de votre entreprise"
    rue_entreprise = "123, Rue de l'Entreprise"
    ville_entreprise = "Ville de l'Entreprise"
    code_postal_entreprise = "12345"

     # Ajouter une photo
    image_path =r"C:\Users\Mega-Pc\OneDrive\Bureau\python\signiature.jpeg "
    c.drawImage(image_path, x=400, y=100, width=2*inch, height=2*inch)




# Ajout des informations de l'entreprise à la facture
    c.setFont("Helvetica", 11)
    c.setFillColorRGB(0, 0, 0)  # Noir
    c.drawString(60, 690, f"Entreprise: {nom_entreprise}")
    c.drawString(60, 670, f"Adresse: {rue_entreprise}, {ville_entreprise}, {code_postal_entreprise}")
    y = 640
    for row in data:
        for x, cell in enumerate(row):
            c.setFont("Helvetica", 9)

            c.drawString(60 + x * 70, y, str(cell))
        y -= 40

    # Enregistrement du PDF
    c.save()
    print(f"La facture a été enregistrée dans {nom_fichier_pdf}")

# Récupération du chemin du dossier contenant le script Python
chemin_dossier_script = os.path.dirname(os.path.abspath(__file__))

# Chargement des données depuis le fichier Excel
chemin_fichier_excel = os.path.join(chemin_dossier_script, 'donneesAchats.xlsx')
wb = load_workbook(chemin_fichier_excel)
ws = wb.active
data = ws.iter_rows(values_only=True)

# Demander à l'utilisateur l'emplacement où enregistrer le fichier PDF
chemin_emplacement_enregistrement = choisir_emplacement_enregistrement()

# Création de la facture à partir des données et enregistrement à l'emplacement choisi
if chemin_emplacement_enregistrement:
    creer_facture(data, chemin_emplacement_enregistrement)
else:
    print("Aucun emplacement sélectionné. La facture n'a pas été enregistrée.")